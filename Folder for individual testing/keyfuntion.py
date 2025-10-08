import pandas as pd
from openpyxl import load_workbook

class LoadCourse:
    def __init__(self, file_path):
        self.file_path = file_path
        self.data = None  # Initialize to avoid errors

    def load_data(self):
        try:
            self.data = pd.read_excel(self.file_path)
            print("Data loaded successfully.")
        except Exception as e:
            print(f"Error loading data: {e}")

    def get_data(self):
        if self.data is not None:
            if 'Course Name' in self.data.columns and 'Course Number' in self.data.columns:
                print("Course Name                Course Number")
                for _, row in self.data.iterrows():
                    print(f"{row['Course Name']}    {row['Course Number']}")
            else:
                print("Required columns 'Course Name' or 'Course Number' not found.")
            return self.data
        else:
            print("No data loaded.")
            return None
        
class LoadJobs:
    def __init__(self, file_path):
        self.file_path = file_path
        self.specializations = {}

    def load_data(self):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            start_row = None

            # Find the row with 'Specialization'
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and str(cell_value).strip().lower() == "specialization":
                    start_row = row + 1
                    break

            if start_row is None:
                print("Keyword 'Specialization' not found.")
                return

            row = start_row
            while row <= sheet.max_row:
                cell = sheet.cell(row=row, column=1)
                value = cell.value

                if value and getattr(cell.font, 'bold', False):
                    specialization = str(value).strip()
                    jobs = []
                    row += 1

                    while row <= sheet.max_row:
                        next_cell = sheet.cell(row=row, column=1)
                        next_value = next_cell.value
                        if not next_value or getattr(next_cell.font, 'bold', False):
                            break
                        jobs.append(str(next_value).strip())
                        row += 1

                    self.specializations[specialization] = jobs
                else:
                    row += 1

            print("Data loaded successfully.")
        except Exception as e:
            print(f"Error loading data: {e}")

    def display_specializations(self):
        for i, (spec, jobs) in enumerate(self.specializations.items(), start=1):
            print(f"\nspecialization_{i}: {spec}")
            print(f"job_{i}:")
            for job in jobs:
                print(f"  - {job}")

if __name__ == "__main__":
    reader = LoadCourse("sources/CPE_Courses.xlsx")
    reader.load_data()
    df = reader.get_data()
    print("\nFull DataFrame:")
    print(df)
    reader_jobs = LoadJobs("sources/CPE_Courses.xlsx")
    reader_jobs.load_data()
    reader_jobs.display_specializations()
