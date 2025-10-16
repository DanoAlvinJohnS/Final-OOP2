import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_PATH = os.path.join("sources", "excels", "user_data.xlsx")

def ensure_excel_exists():
    """Creates the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_PATH):
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["User ID", "Username", "Password", "Email", "Date Created"])
        wb.save(EXCEL_PATH)

def get_next_user_id():
    """Reads the last User ID and returns the next available one."""
    ensure_excel_exists()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ids = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            ids.append(int(row[0]))

    wb.close()
    if not ids:
        return "0001"
    return str(max(ids) + 1).zfill(4)

def save_user(username, password, email):
    """Saves a new user into the Excel file."""
    ensure_excel_exists()
    user_id = get_next_user_id()
    date_created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([user_id, username, password, email, date_created])
    wb.save(EXCEL_PATH)
    wb.close()

    return user_id

def get_all_users():
    if not os.path.exists(EXCEL_PATH):
        print("Excel file not found at:", EXCEL_PATH)
        return []

    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active

    users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip headers
        user_id, username, password, email, date_created = row
        if username:  # skip empty rows
            users.append({
                "user_id": str(user_id),
                "username": username.strip(),
                "password": password.strip(),
                "email": email.strip(),
                "date_created": date_created
            })
    wb.close()

    users.sort(key=lambda x: x["username"].lower())
    return users

def binary_search_user(users, target_username):
    low = 0
    high = len(users) - 1
    target = target_username.lower()

    while low <= high:
        mid = (low + high) // 2
        mid_username = users[mid]["username"].lower()

        if mid_username == target:
            return users[mid]
        elif mid_username < target:
            low = mid + 1
        else:
            high = mid - 1

    return None  


